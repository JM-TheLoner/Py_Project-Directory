from tkinter import *  # type: ignore
from tkinter import messagebox
import random
from PIL import ImageTk, Image
import openpyxl as xl


wb = xl.load_workbook("Py_Project Directory\\bank_db.xlsx") 
ws_1 = wb['Sheet1']

Allowed = "1234567890"


def logpage():
    new = Tk()
    root.destroy()
    user = name.get()
    nows=0
    row_num=0
    for i in range(2, ws_1.max_row+1):
        bs = ws_1.cell(row=i, column=3)
        bsv= bs.value
        bsvs = str(bsv)
        if user != bsvs:
            nows+=1
        else:
            row_num+=i
            break
    
    l = ws_1.cell(row=row_num,column=5)
    f = ws_1.cell(row=row_num,column=4)
    ln=l.value
    fn=f.value
    no = ws_1.cell(row=row_num,column=1)
    mo = no.value


    new.geometry("1000x680")
    new.title(f"Welcome")
    new.resizable(False,False)
    new.config(bg="black")

    labels = Label(new, text=f"Welcome, {ln} {fn}\n", width=30, font=("arial", 39, "bold"),  bg="black", fg="red")
    labels.pack()

    labels = Label(new, text=f"{mo}", width=15, font=("arial", 20, "bold"),  bg="black", fg="red")
    labels.place(x=350, y=100)

    labels = Label(new, text="Choose an action", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
    labels.place(x=300, y=150) 


    def check():
        bal = ws_1.cell(row=row_num, column=8)
        balance = bal.value

        box = Tk()
        box.overrideredirect()
        box.withdraw()
        messagebox.showinfo("Balance", f"Account Balance: ${balance}")
        box.destroy()                


    def loan():
        amm = ws_1.cell(row=row_num, column=8)
        ammo = amm.value
        allowance = int(ammo) * 0.1 #type: ignore

        l = ws_1.cell(row=row_num, column=12)
        ll = l.value   
        lll = int(ll) # type: ignore
        c = 0

        if lll == 0:
            ll = "-"
        else:
            c = 1

        u = ws_1.cell(row=row_num, column=11)
        uu = u.value   
        uuu = int(uu) # type: ignore
        c = 0

        if uuu == 0:
          uu = "-"
        else:
            c = 1

        l = Tk()
        l.geometry("1000x500")
        l.title(f"Take a loan")
        l.resizable(False,False)
        l.config(bg="black")

        labels = Label(l, text=f"Max loan available to you: ${ll}\ncurrent debt: ${uu}", width=40, font=("arial", 20, "bold"),  bg="black", fg="red")
        labels.place(x=130, y=90)

        labels = Label(l, text="Input amount", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
        labels.place(x=300, y=160) 

        labels = Label(l, text=f"Max loan at this time: ${allowance}", width=40, font=("arial", 15, "bold"),  bg="black", fg="red")
        labels.place(x=250, y=210)

        labels = Label(l, text="Take a Loan", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
        labels.pack()

        num=IntVar()
        Entry(width=15, font=("arial", 28), textvariable=num, bd=0, border=2).place(x=320, y=240)
        
        labels = Label(l, text="Enter Passkey", width=15, font=("arial", 30, "bold"),  bg="black", fg="red").place(x=300, y=300)

        nums=StringVar()
        Entry(width=15, font=("arial", 28), textvariable=nums, bd=0, border=2, show="*").place(x=320, y=350)

        def doit():
            ask = num.get()
            pin = (nums.get())

            pas = ws_1.cell(row=row_num, column=2)
            passes = str(pas.value)
            amm = ws_1.cell(row=row_num, column=8)
            ammo = amm.value
            maxo = ws_1.cell(row=row_num, column=9)
            maxou = maxo.value
            otherd = int(maxou) * 1.15 # type: ignore
            newval = otherd - int(ammo) # type: ignore

            allowed = int(ammo) * 0.1  # type: ignore

            deb = ws_1.cell(row=row_num, column=11)
            debb = deb.value
            debt = int(debb)  # type: ignore
            maxes = debt + ask  # type: ignore

            lim = ws_1.cell(row=row_num, column=12)# type: ignore
            limi = lim.value    
            limit = int(limi)     # type: ignore    

            debtor = debt

            if int(pin) == int(passes):
                if limit == 0:
                    limit += newval
                    lim.value = limit
                
                if maxes < limit:
                    if ask <= allowed:
                        new_amt = int(ammo) + ask  # type: ignore
                        amm.value = (new_amt)

                        maxi = ws_1.max_row

                        admin = ws_1.cell(row=maxi, column=8)
                        admins = admin.value

                        adminis = int(admins) - ask  # type: ignore
                        admin.value = adminis

                        debtor= debt + ask

                        deb.value=debtor

                        wb.save("Py_Project Directory\\bank_db.xlsx")

                        box = Tk()
                        box.overrideredirect()
                        box.withdraw()
                        messagebox.showinfo("Take a Loan", f"Loan Successful")
                        box.destroy()    

                        l.destroy()

                    else:
                        box = Tk()
                        box.overrideredirect()
                        box.withdraw()
                        messagebox.showinfo("Loan Error", f"amount exceeds allowable loan")
                        box.destroy() 

                        l.destroy()
                
                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Loan Error", f"Repay some of the debt you owe\nYou owe ${debtor}")
                    box.destroy() 

                    l.destroy()
            
            else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Error", f"Wrong Passkey")
                    box.destroy() 

                    l.destroy()          

        
        Button(l, width=20, text="Take Loan", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=doit).place(x=280 ,y=420)


    def withdraw():
        
        a = Tk()

        a.geometry("1000x600")
        a.title(f"Welcome, {ln} {fn}")
        a.resizable(False,False)
        a.config(bg="black")

        labels = Label(a, text="WITHDRAWAL", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
        labels.pack()

        labels = Label(a, text="Choose an amount", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
        labels.place(x=300, y=150) 

        amm = ws_1.cell(row=row_num, column=8)# type: ignore
        ammo = amm.value
        ico = ws_1.cell(row=row_num, column=9)# type: ignore
        icol = ico.value
        icols = int(icol)  # type: ignore

        def five_h():
            if int(ammo) >= 500: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 500
                amm.value = (new_amt)
                icolss = icols - 500
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()
            
            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()

                a.destroy()


        def one_k():
            if int(ammo) >= 1000: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 1000
                amm.value = (new_amt)
                icolss = icols - 1000
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()

            
            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()

                a.destroy()

        def two_k():
            if int(ammo) >= 2000: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 2000
                amm.value = (new_amt)
                icolss = icols - 2000
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()
            
            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()

                a.destroy()

        def five_k():
            if int(ammo) >= 5000: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 5000
                amm.value = (new_amt)
                icolss = icols - 5000
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()
            
            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()
                
                a.destroy()

        def ten_k():
            if int(ammo) >= 10000: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 10000
                amm.value = (new_amt)
                icolss = icols - 1000
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()

            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()

                a.destroy()

        def fifteen_k():
            if int(ammo) >= 15000: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 15000
                amm.value = (new_amt)
                icolss = icols - 15000
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()

            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()

                a.destroy()

        def twenty_k():
            if int(ammo) >= 20000: # type: ignore
                amt = int(ammo) # type: ignore
                new_amt = amt - 20000
                amm.value = (new_amt)
                icolss = icols - 20000
                ico.value = (icolss)

                wb.save("Py_Project Directory\\bank_db.xlsx")

                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Successful","Withdrawal Complete")
                box.destroy() 

                a.destroy()
            
            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo(f"Error","Insufficient Funds!!!")
                box.destroy()

                a.destroy()

        def other():
            a.destroy()
            other = Tk()
            other.geometry("1000x400")
            other.title(f"Welcome, {ln} {fn}")
            other.resizable(False,False)
            other.config(bg="black")

            labels = Label(other, text="WITHDRAWAL", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(other, text="Input amount", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100) 

            num=IntVar()
            Entry(other, width=15, font=("arial", 28), textvariable=num, bd=0, border=2).place(x=320, y=230)

            def others():
                a = num.get()
                if int(ammo) >= a: # type: ignore
                    amt = int(ammo) # type: ignore
                    new_amt = amt - a
                    amm.value = (new_amt)
                    icolss = icols - a
                    ico.value = (icolss)

                    wb.save("Py_Project Directory\\bank_db.xlsx")

                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo(f"Successful","Withdrawal Complete")
                    box.destroy() 

                    other.destroy()

                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                    box.destroy()

                    other.destroy()
            
            
            Button(other, width=20, text="Withdraw", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=others).place(x=280 ,y=300)


        Button(a, width=20, text="500", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=five_h).place(x=20 ,y=230)
        Button(a, width=20, text="1000", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=one_k).place(x=20 ,y=320)
        Button(a, width=20, text="2000", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=two_k).place(x=20 ,y=410)
        Button(a, width=20, text="10000", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=ten_k).place(x=550 ,y=230)
        Button(a, width=20, text="15000", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=fifteen_k).place(x=550 ,y=320)
        Button(a, width=20, text="20000", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=twenty_k).place(x=550 ,y=410)
        Button(a, width=20, text="5000", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=five_k).place(x=20 ,y=500)
        Button(a, width=20, text="Other Amount", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=other).place(x=550 ,y=500)


    def change():
            change = Tk()
            change.geometry("1000x500")
            change.title(f"Welcome, {name.get()}")
            change.resizable(False,False)
            change.config(bg="black")

            labels = Label(change, text="Change Pin", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(change, text="Input old pin", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100)

            nums=StringVar()
            Entry(width=15, font=("arial", 28), textvariable=nums, bd=0, border=2, show="*").place(x=320, y=180)

            labels = Label(change, text="Input new pin", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=260)

            num=StringVar()
            Entry(width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=340)       
            

            def do():
                olds = nums.get()
                pas = ws_1.cell(row=row_num, column=2)
                passes = str(pas.value)

                if olds == passes:
                    pins = num.get()
                    pas.value = int(pins)

                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Change Pin", f"Pin Changed Successfully")
                    box.destroy()

                    wb.save("Py_Project Directory\\bank_db.xlsx")

                    change.destroy()
                
                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Change Pin Error", f"Please input the correct old pin")
                    box.destroy()        
            

            Button(change, width=20, text="Change Pin", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=do).place(x=280 ,y=420)
   

    def transfer():
            tran = Tk()
            tran.geometry("1000x500")
            tran.title(f"Welcome, {name.get()}")
            tran.resizable(False,False)
            tran.config(bg="black")

            labels = Label(tran, text="transfer", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(tran, text="Destination Account", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100)

            nums=StringVar()
            Entry(width=15, font=("arial", 28), textvariable=nums, bd=0, border=2).place(x=320, y=180)

            labels = Label(tran, text="Input Amount", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=260)

            num=StringVar()
            Entry(width=15, font=("arial", 28), textvariable=num, bd=0, border=2).place(x=320, y=340)    

            pas = ws_1.cell(row=row_num, column=2)
            passes = str(pas.value)
            amm = ws_1.cell(row=row_num, column=8)
            ammo = amm.value 
            userd = ws_1.cell(row=row_num, column=1)
            users = userd.value
            bank = ws_1.cell(row=row_num, column=10)
            banks = bank.value
            ico = ws_1.cell(row=row_num, column=9)
            icol = ico.value
            
            def proceed():
                userf=nums.get()
                nows=0
                row_numa=0
                userb_name=""
                for i in range(2, ws_1.max_row+1):
                    bsa = ws_1.cell(row=i, column=1)
                    bsva= bsa.value
                    bsvsa = str(bsva)
                    if userf != bsvsa:
                        nows+=1
                    else:
                            row_numa+=i
                            userb_na = ws_1.cell(row=row_numa, column=1)
                            userb_name = userb_na.value

                            la = ws_1.cell(row=row_numa,column=5)
                            fa = ws_1.cell(row=row_numa,column=4)
                            lan=la.value
                            fan=fa.value
                            ammot = ws_1.cell(row=row_numa, column=8)
                            ammoa = ammot.value
                            banka = ws_1.cell(row=row_numa, column=10)
                            banksa = banka.value
                            icoa = ws_1.cell(row=row_numa, column=9)
                            icola = icoa.value

                            break    
                
                if nows != ws_1.max_row-1:
                    if userb_name != users:     # type: ignore
                                tran.destroy()
                                tr= Tk()
                                tr.geometry("1000x360")
                                tr.title(f"Transfer")
                                tr.resizable(False,False)
                                tr.config(bg="black")  

                                labels = Label(tr, text=f"reciever: {lan} {fan}", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")   # type: ignore
                                labels.pack()

                                labels = Label(tr, text="Enter Passkey", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
                                labels.place(x=300, y=100) 

                                numu=StringVar()
                                Entry(tr, width=15, font=("arial", 28), textvariable=numu, bd=0, border=2, show="*").place(x=320, y=200)

                                def conf():
                                    trial=numu.get()
                                    req=num.get()

                                    if int(req) <= int(ammo): #type: ignore
                                        if str(trial) == str(passes):
                                            amta = int(ammo)    # type: ignore
                                            amtb = int(ammoa)    # type: ignore
                                            icolsa = int(icola)  # type: ignore
                                            icols = int(icol)  # type: ignore  

                                            icolsa += int(req)
                                            amta -= int(req)
                                            amtb += int(req)
                                            icols -= int(req)

                                            amm.value = (amta)
                                            ammot.value = (amtb)  
                                            icoa.value = (icolsa)
                                            ico.value = (icols)

                                            if banks != banksa:
                                                levy = int(req)*0.01

                                                maxi = ws_1.max_row

                                                admin = ws_1.cell(row=maxi, column=8)
                                                admins = admin.value
                                                adminis = int(admins) + levy #type: ignore
                                                admin.value = adminis

                                                amts= amta - (levy)
                                                amm.value = amts
                                                ands = icols - (levy)
                                                ico.value = ands

                                            else:
                                                levy = 0
                                                                                        
                                            wb.save("Py_Project Directory\\bank_db.xlsx")

                                            box = Tk()
                                            box.overrideredirect()
                                            box.withdraw()
                                            messagebox.showinfo(f"Successful","Transfer Complete")
                                            box.destroy() 

                                            tr.destroy()  

                                        else:
                                            box = Tk()
                                            box.overrideredirect()
                                            box.withdraw()
                                            messagebox.showinfo(f"Error","Wrong passkey")
                                            box.destroy() 

                                            tr.destroy()
                                    
                                    else:
                                        box = Tk()
                                        box.overrideredirect()
                                        box.withdraw()
                                        messagebox.showinfo(f"Error","You don't have enough money")
                                        box.destroy() 

                                        tr.destroy()                               
                                        
                                Button(tr, width=20, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=conf).place(x=280 ,y=280)

                    else: 
                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Error", f"Did you just try to transfer money to yourself?\nBloddy Idiot")
                            box.destroy()                           

                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Error", f"Account number inaccurate\nNo warning next time... Your Money will just vanish\n\n\nJust know that there is no bank branch to run to")
                    box.destroy()                       

            Button(tran, width=20, text="Proceed", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=proceed).place(x=280 ,y=420)
 

    def logout():
        new.destroy()
    

    def terminate():
        state = Tk()

        state.geometry("600x280")
        state.title(f"Account Termination")
        state.resizable(False,False)
        state.config(bg="black")

        labels = Label(state, text="Terminate Account", width=15, font=("arial", 35, "bold"),  bg="black", fg="red")
        labels.pack()

        labels = Label(state, text="Input Passkey", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
        labels.place(x=140, y=80)

        num=StringVar()
        Entry(width=15, font=("arial", 28), textvariable=num, bd=0, border=2,show="*").place(x=150, y=150) 

        def term():
            fine=num.get()
            pas = ws_1.cell(row=row_num, column=2)
            passes = str(pas.value)
            if fine == passes:
                last = Tk()

                last.geometry("500x280")
                last.title(f"Account Termination")
                last.resizable(False,False)
                last.config(bg="black")

                labels = Label(last, text="Are You Sure", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                labels.pack()

                def yes():
                    a=ws_1.cell(row=row_num, column=1)
                    b=ws_1.cell(row=row_num, column=2)
                    c=ws_1.cell(row=row_num, column=3)
                    d=ws_1.cell(row=row_num, column=4)
                    e=ws_1.cell(row=row_num, column=5)
                    f=ws_1.cell(row=row_num, column=6)
                    g=ws_1.cell(row=row_num, column=7)
                    h=ws_1.cell(row=row_num, column=8)
                    i=ws_1.cell(row=row_num, column=9)
                    j=ws_1.cell(row=row_num, column=10)

                    a.value = ""
                    b.value = ""
                    c.value = ""
                    d.value = ""
                    e.value = ""
                    f.value = ""
                    g.value = ""
                    h.value = ""
                    i.value = ""
                    j.value = ""
                    
                    wb.save("Py_Project Directory\\bank_db.xlsx")

                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Termination Notice", "Your Account has been Terminated\n Thank you for Banking with us")
                    box.destroy()
                    
                    last.destroy()
                    new.destroy()
                    state.destroy()
                    

                def no():
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Termination cancelled", "You came so far\nOnly to turn tail like the Bitch you are\nGood to know we own you")
                    box.destroy()
                    last.destroy()   

                Button(last, width=9, text="YES", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=yes).place(x=20 ,y=160)
                Button(last, width=9, text="NO", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=no).place(x=270 ,y=160)


            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo("Termination Error", "Wrong passkey\nImagine failing suicide")
                box.destroy()
                state.destroy()
                

        Button(state, width=12, text="Terminate", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=term).place(x=185 ,y=210)


    def repay():
            lef = ws_1.cell(row=row_num,column=8)  
            left = lef.value
            owe = ws_1.cell(row=row_num,column=11)  
            ower = owe.value
            owing = int(ower) # type: ignore
            p = ws_1.cell(row=row_num,column=2)
            pw = p.value
            lim = ws_1.cell(row=row_num,column=12) 
            limi = lim.value
            m = ws_1.max_row 
            adm = ws_1.cell(row=m,column=8)
            admin = adm.value
        
            if owing != 0:
                safe = Tk()
                safe.geometry("1000x480")
                safe.title(f"Repay Debt")
                safe.resizable(False,False)
                safe.config(bg="black")   

                labels = Label(safe, text=f"Repay Debt", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                labels.pack()

                labels = Label(safe, text=f"You owe ${ower}\nInput Amount to Pay", width=30, font=("arial", 28, "bold"),  bg="black", fg="red")
                labels.place(x=110, y=100) 

                phone=StringVar()
                Entry(safe, width=15, font=("arial", 28), textvariable=phone, bd=0, border=2).place(x=320, y=200) 

                labels = Label(safe, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                labels.place(x=300, y=270) 

                amount=StringVar()
                Entry(safe, width=15, font=("arial", 28), textvariable=amount, bd=0, border=2, show="*").place(x=320, y=320)


                def repays():
                        debt = phone.get()
                        debtt = int(debt)
                        pin = amount.get()
                        pis = str(pin)
        
                        if int(left) >= debtt:  # type: ignore
                            if pis == str(pw):
                                if debtt <= owing:
                                    remain = int(left) - debtt # type: ignore
                                    owes = int(ower) - debtt  # type: ignore
                                    good = int(admin) + debtt  # type: ignore  

                                    lef.value = remain
                                    owe.value = owes
                                    adm.value = good

                                    wb.save("Py_Project Directory\\bank_db.xlsx")
                                    

                                    if owe.value == 0:
                                        lim.value = 0 # type: ignore
                                        box = Tk()
                                        box.overrideredirect()
                                        box.withdraw()
                                        messagebox.showinfo("Debt Settlement", f"All Debt Successfully Setteled")
                                        box.destroy() 

                                        wb.save("Py_Project Directory\\bank_db.xlsx") 
                                        safe.destroy()

                                    else: 
                                        box = Tk()
                                        box.overrideredirect()
                                        box.withdraw()
                                        messagebox.showinfo("Debt Settlement", f"Payment Successful\nYou still have ${owes} outstanding")
                                        box.destroy()

                                        wb.save("Py_Project Directory\\bank_db.xlsx")
                                        safe.destroy()

                                else: 
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo("Debt Settlement", f"Amount entered is greater than debt owed")
                                    box.destroy()

                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo("Debt Settlement", f"Incorrect Passkey")
                                box.destroy() 
                        
                        else:
                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Debt Settlement", f"Insufficient Funds")
                            box.destroy()                             
            
                Button(safe, width=12, text="Repay debt", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=repays).place(x=350 ,y=390)
          
            else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Debt Settlement", f"You have no Debt")
                    box.destroy() 


    def statement():
        state = Tk()

        state.geometry("1000x400")
        state.title(f"Account Statement")
        state.resizable(False,False)
        state.config(bg="black")

        labels = Label(state, text="Account Statement", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
        labels.pack()

        labels = Label(state, text="Input Email", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
        labels.place(x=300, y=100)

        num=StringVar()
        Entry(width=15, font=("arial", 28), textvariable=num, bd=0, border=2).place(x=320, y=230) 

        def does():
            ema = ws_1.cell(row=row_num,column=6)  
            email = ema.value 
            mail = num.get()
            numb = ws_1.cell(row=row_num,column=1)  
            number = numb.value

            if email == mail:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo("Account Statement", f"Statement for account: {number} Will Be Sent To {mail} shortly")
                box.destroy() 

                state.destroy()

            else:
                box = Tk()
                box.overrideredirect()
                box.withdraw()
                messagebox.showinfo("Account Statement", f"Wrong Email Address Provided")
                box.destroy()         

                state.destroy()       

        Button(state, width=20, text="Send Statement", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=does).place(x=280 ,y=300)

    def quickteller():
        qt = Tk()

        qt.geometry("1000x360")
        qt.title(f"Quick Teller")
        qt.resizable(False,False)
        qt.config(bg="black")

        labels = Label(qt, text=f"Quick Teller", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
        labels.pack()

        labels = Label(qt, text="Choose an action", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
        labels.place(x=300, y=100) 

        def safe():
            qt.destroy()
            safe = Tk()
        
            safe.geometry("1000x360")
            safe.title(f"Quick Teller")
            safe.resizable(False,False)
            safe.config(bg="black")   

            labels = Label(safe, text=f"Generate OTP", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(safe, text="Enter Passkey", width=15, font=("arial", 30, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100) 

            num=StringVar()
            Entry(safe, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=200) 

            def safer():
                gim = num.get()
                pas = ws_1.cell(row=row_num, column=2)
                passes = str(pas.value)

                if gim == passes:
                    import random  #important                           #not that difficult though
                    length = 4
                    otp = "".join(random.sample(Allowed,length))

                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Generate OTP", f"Your OTP is {otp}\nDo Not Share This OTP with anyone")
                    box.destroy()    

                    safe.destroy()

                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Passkey Error", f"Password provided is wrong")
                    box.destroy()    

                    safe.destroy() 

            Button(safe, width=20, text="Generate", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=safer).place(x=280 ,y=280)


        def airtime():
            qt.destroy()
            safe = Tk()
        
            safe.geometry("1000x430")
            safe.title(f"Quick Teller")
            safe.resizable(False,False)
            safe.config(bg="black")   

            labels = Label(safe, text=f"Buy Airtime", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(safe, text="Phone Number", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100) 

            phone=StringVar()
            Entry(safe, width=15, font=("arial", 28), textvariable=phone, bd=0, border=2).place(x=320, y=150) 

            labels = Label(safe, text="Enter Amount", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=220) 

            amount=StringVar()
            Entry(safe, width=15, font=("arial", 28), textvariable=amount, bd=0, border=2).place(x=320, y=270) 


            def buyit():
                ph = (phone.get())


                if len(ph) == 11:
                    safe.destroy()
                    safer = Tk()
            
                    safer.geometry("1000x360")
                    safer.title(f"Quick Teller")
                    safer.resizable(False,False)
                    safer.config(bg="black")   

                    labels = Label(safer, text=f"Buy Airtime", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                    labels.pack()

                    labels = Label(safer, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                    labels.place(x=300, y=100) 

                    num=StringVar()
                    Entry(safer, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Error", f"You have Entered a wrong phone number")
                    box.destroy() 


                def right():
                    want = int(amount.get())
                    pin = num.get()
                    pas = ws_1.cell(row=row_num, column=2)
                    passes = str(pas.value)
                    amm = ws_1.cell(row=row_num, column=8)
                    ammo = amm.value
                    ico = ws_1.cell(row=row_num, column=9)
                    icol = ico.value

                    if pin == passes:
                        if int(want) <= int(ammo):  # type: ignore
                            amt = ammo
                            new_amt = int(amt) - want  # type: ignore
                            amm.value = new_amt
                            icols = icol
                            icolss = int(icols) - want  # type: ignore
                            ico.value = icolss

                            wb.save("Py_Project Directory\\bank_db.xlsx")

                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Purchase Successful", f"Purchase Successful ")
                            box.destroy() 

                        else:
                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Purchase Unsuccessful", f"INSUFFICIENT FUNDS")
                            box.destroy()

                    else:
                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Error", f"You have entered the wrong passkey")
                            box.destroy()

                    safer.destroy()

                Button(safer, width=20, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=right).place(x=280 ,y=280)


            Button(safe, width=20, text="Proceed", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=buyit).place(x=280 ,y=350)

        def dstv():
            qt.destroy()
            safe = Tk()
        
            safe.geometry("1000x300")
            safe.title(f"Quick Teller")
            safe.resizable(False,False)
            safe.config(bg="black")   

            labels = Label(safe, text=f"Subscribe to DSTV", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(safe, text="Smartcard Number", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100) 

            phone=StringVar()
            Entry(safe, width=18, font=("arial", 28), textvariable=phone, bd=0, border=2).place(x=300, y=150) 

            def buyit():
                ph = (phone.get())

                if len(ph) == 16:
                    safe.destroy()
                    safer = Tk()
            
                    safer.geometry("1000x550")
                    safer.title(f"Quick Teller")
                    safer.resizable(False,False)
                    safer.config(bg="black")   

                    labels = Label(safer, text=f"Subscribe to Dstv", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                    labels.pack()

                    labels = Label(safer, text="Choose Plan", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                    labels.place(x=300, y=100) 

                    pas = ws_1.cell(row=row_num, column=2)# type: ignore
                    passes = str(pas.value)
                    amm = ws_1.cell(row=row_num, column=8)# type: ignore
                    ammo = amm.value
                    ico = ws_1.cell(row=row_num, column=9)# type: ignore
                    icol = ico.value

                    def padi():
                        safer.destroy()
                        safes = Tk()
                
                        safes.geometry("1000x360")
                        safes.title(f"Quick Teller")
                        safes.resizable(False,False)
                        safes.config(bg="black")   

                        labels = Label(safes, text=f"Dstv Padi", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                        labels.pack()

                        labels = Label(safes, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                        labels.place(x=300, y=100) 

                        num=StringVar()
                        Entry(safes, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                        def final():
                            pin = num.get()

                            if (pin) == passes:
                                if int(ammo) >= 2500:   # type: ignore
                                    amt = int(ammo)  # type: ignore
                                    new_amt = amt - 2500
                                    amm.value = (new_amt)
                                    icols = icol
                                    icolss = int(icols) - 2500  # type: ignore
                                    ico.value = icolss


                                    wb.save("Py_Project Directory\\bank_db.xlsx")

                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Successful","Subscription Successful")
                                    box.destroy()                                     
                                
                                else:
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                                    box.destroy()
                                
                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo(f"Error","Wrong Passkey")
                                box.destroy()

                            safes.destroy()

                                
                        Button(safes, width=17, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=final).place(x=320 ,y=230)

                    def yanga():
                        safer.destroy()
                        safes = Tk()
                
                        safes.geometry("1000x360")
                        safes.title(f"Quick Teller")
                        safes.resizable(False,False)
                        safes.config(bg="black")   

                        labels = Label(safes, text=f"Dstv Yanga", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                        labels.pack()

                        labels = Label(safes, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                        labels.place(x=300, y=100) 

                        num=StringVar()
                        Entry(safes, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                        def final():
                            pin = num.get()

                            if (pin) == passes:
                                if int(ammo) >= 3500:   # type: ignore
                                    amt = int(ammo)  # type: ignore
                                    new_amt = amt - 3500
                                    amm.value = (new_amt)
                                    icols = icol
                                    icolss = int(icols) - 3500  # type: ignore
                                    ico.value = icolss


                                    wb.save("Py_Project Directory\\bank_db.xlsx")

                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Successful","Subscription Successful")
                                    box.destroy() 
                                
                                else:
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                                    box.destroy()
                                
                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo(f"Error","Wrong Passkey")
                                box.destroy()

                            safes.destroy()

                        Button(safes, width=17, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=final).place(x=320 ,y=230)                    

                    def confam():
                        safer.destroy()
                        safes = Tk()
                
                        safes.geometry("1000x360")
                        safes.title(f"Quick Teller")
                        safes.resizable(False,False)
                        safes.config(bg="black")   

                        labels = Label(safes, text=f"Dstv Confam", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                        labels.pack()

                        labels = Label(safes, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                        labels.place(x=300, y=100) 

                        num=StringVar()
                        Entry(safes, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                        def final():
                            pin = num.get()

                            if (pin) == passes:
                                if int(ammo) >= 6200:   # type: ignore
                                    amt = int(ammo)  # type: ignore
                                    new_amt = amt - 6200
                                    amm.value = (new_amt)
                                    icols = icol
                                    icolss = int(icols) - 6200  # type: ignore
                                    ico.value = icolss

                                    wb.save("Py_Project Directory\\bank_db.xlsx")

                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Successful","Subscription Successful")
                                    box.destroy() 
                                
                                else:
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                                    box.destroy()
                                
                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo(f"Error","Wrong Passkey")
                                box.destroy()

                            safes.destroy()


                        Button(safes, width=17, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=final).place(x=320 ,y=230)

                    def comp():
                        safer.destroy()
                        safes = Tk()
                
                        safes.geometry("1000x360")
                        safes.title(f"Quick Teller")
                        safes.resizable(False,False)
                        safes.config(bg="black")   

                        labels = Label(safes, text=f"Dstv Compact", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                        labels.pack()

                        labels = Label(safes, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                        labels.place(x=300, y=100) 

                        num=StringVar()
                        Entry(safes, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                        def final():
                            pin = num.get()

                            if (pin) == passes:
                                if int(ammo) >= 10500:   # type: ignore
                                    amt = int(ammo)  # type: ignore
                                    new_amt = amt - 10500
                                    amm.value = (new_amt)
                                    icols = icol
                                    icolss = int(icols) - 10500  # type: ignore
                                    ico.value = icolss

                                    wb.save("Py_Project Directory\\bank_db.xlsx")

                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Successful","Subscription Successful")
                                    box.destroy() 
                                
                                else:
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                                    box.destroy()
                                
                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo(f"Error","Wrong Passkey")
                                box.destroy()

                            safes.destroy()


                        Button(safes, width=17, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=final).place(x=320 ,y=230)

                    def compP():
                        safer.destroy()
                        safes = Tk()
                
                        safes.geometry("1000x360")
                        safes.title(f"Quick Teller")
                        safes.resizable(False,False)
                        safes.config(bg="black")   

                        labels = Label(safes, text=f"Dstv Compact Plus", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                        labels.pack()

                        labels = Label(safes, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                        labels.place(x=300, y=100) 

                        num=StringVar()
                        Entry(safes, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 


                        def final():
                            pin = num.get()

                            if (pin) == passes:
                                if int(ammo) >= 16600:   # type: ignore
                                    amt = int(ammo)  # type: ignore
                                    new_amt = amt - 16600
                                    amm.value = (new_amt)
                                    icols = icol
                                    icolss = int(icols) - 16600  # type: ignore
                                    ico.value = icolss

                                    wb.save("Py_Project Directory\\bank_db.xlsx")

                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Successful","Subscription Successful")
                                    box.destroy() 
                                
                                else:
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                                    box.destroy()
                                
                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo(f"Error","Wrong Passkey")
                                box.destroy()

                            safes.destroy()


                        Button(safes, width=17, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=final).place(x=320 ,y=230)

                    def prem():
                        safer.destroy()
                        safes = Tk()
                
                        safes.geometry("1000x360")
                        safes.title(f"Quick Teller")
                        safes.resizable(False,False)
                        safes.config(bg="black")   

                        labels = Label(safes, text=f"Dstv Premium", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                        labels.pack()

                        labels = Label(safes, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                        labels.place(x=300, y=100) 

                        num=StringVar()
                        Entry(safes, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                        def final():
                            pin = num.get()

                            if (pin) == passes:
                                if int(ammo) >= 24500:   # type: ignore
                                    amt = int(ammo)  # type: ignore
                                    new_amt = amt - 24500
                                    amm.value = (new_amt)
                                    icols = icol
                                    icolss = int(icols) - 24500  # type: ignore
                                    ico.value = icolss

                                    wb.save("Py_Project Directory\\bank_db.xlsx")

                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Successful","Subscription Successful")
                                    box.destroy() 
                                
                                else:
                                    box = Tk()
                                    box.overrideredirect()
                                    box.withdraw()
                                    messagebox.showinfo(f"Error","Insufficient Funds!!!")
                                    box.destroy()
                                
                            else:
                                box = Tk()
                                box.overrideredirect()
                                box.withdraw()
                                messagebox.showinfo(f"Error","Wrong Passkey")
                                box.destroy()

                            safes.destroy()

                        Button(safes, width=17, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=final).place(x=320 ,y=230)

                    Button(safer, width=17, text="Dstv Padi \n $2500", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=padi).place(x=80 ,y=160)
                    Button(safer, width=17, text="Dstv Yanga \n $3500", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=yanga).place(x=80 ,y=280)
                    Button(safer, width=17, text="Dstv Confam \n $6200", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=confam).place(x=80 ,y=400)
                    Button(safer, width=17, text="Dstv Compact \n $10500", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=comp).place(x=550 ,y=160)
                    Button(safer, width=17, text="Dstv Compact Plus \n $16600", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=compP).place(x=550 ,y=280)
                    Button(safer, width=17, text="Dstv Premium \n $24500", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=prem).place(x=550 ,y=400)

                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Error", f"You have Entered a Wrong DSTV Smartcard Number")
                    box.destroy() 


            Button(safe, width=20, text="Proceed", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=buyit).place(x=280 ,y=220)


        def bet():
            qt.destroy()
            safe = Tk()
        
            safe.geometry("1000x430")
            safe.title(f"Quick Teller")
            safe.resizable(False,False)
            safe.config(bg="black")   

            labels = Label(safe, text=f"Transfer to your bet9ja account", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
            labels.pack()

            labels = Label(safe, text="Account Number", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=100) 

            phone=StringVar()
            Entry(safe, width=15, font=("arial", 28), textvariable=phone, bd=0, border=2).place(x=320, y=150) 

            labels = Label(safe, text="Enter Amount", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
            labels.place(x=300, y=220) 

            amount=StringVar()
            Entry(safe, width=15, font=("arial", 28), textvariable=amount, bd=0, border=2).place(x=320, y=270) 


            def buyit():
                ph = (phone.get())

                if len(ph) == 10:
                    safe.destroy()
                    safer = Tk()
            
                    safer.geometry("1000x360")
                    safer.title(f"Quick Teller")
                    safer.resizable(False,False)
                    safer.config(bg="black")   

                    labels = Label(safer, text=f"Transfer", width=15, font=("arial", 50, "bold"),  bg="black", fg="red")
                    labels.pack()

                    labels = Label(safer, text="Enter Passkey", width=15, font=("arial", 28, "bold"),  bg="black", fg="red")
                    labels.place(x=300, y=100) 

                    num=StringVar()
                    Entry(safer, width=15, font=("arial", 28), textvariable=num, bd=0, border=2, show="*").place(x=320, y=150) 

                else:
                    box = Tk()
                    box.overrideredirect()
                    box.withdraw()
                    messagebox.showinfo("Error", f"You have Entered a wrong account number")
                    box.destroy() 


                def right():
                    want = int(amount.get())
                    pin = num.get()

                    pas = ws_1.cell(row=row_num, column=2)
                    passes = str(pas.value)
                    amm = ws_1.cell(row=row_num, column=8)
                    ammo = amm.value
                    ico = ws_1.cell(row=row_num, column=9)# type: ignore
                    icol = ico.value

                    if pin == passes:
                        if int(want) <= ammo: # type: ignore
                            amt = ammo
                            new_amt = int(amt) - want  # type: ignore
                            amm.value = (new_amt)                            
                            icols = icol
                            icolss = int(icols) - want  # type: ignore
                            ico.value = icolss

                            wb.save("Py_Project Directory\\bank_db.xlsx")

                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Transfer Successful", f"Transfer Successful ")
                            box.destroy() 

                        else:
                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Purchase Unsuccessful", f"INSUFFICIENT FUNDS")
                            box.destroy()

                    else:
                            box = Tk()
                            box.overrideredirect()
                            box.withdraw()
                            messagebox.showinfo("Error", f"You have entered the wrong passkey")
                            box.destroy()

                    safer.destroy()

                Button(safer, width=20, text="Confirm", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=right).place(x=280 ,y=280)


            Button(safe, width=20, text="Proceed", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=buyit).place(x=280 ,y=350)


        Button(qt, width=20, text="Airtime", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=airtime).place(x=20 ,y=180)
        Button(qt, width=20, text="Dstv", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=dstv).place(x=20 ,y=280)
        Button(qt, width=20, text="Safe Tokens", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=safe).place(x=550 ,y=180)
        Button(qt, width=20, text="Bet9ja", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=bet).place(x=550 ,y=280)
        

    Button(new, width=20, text="Check Balance", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=check).place(x=20 ,y=230)
    Button(new, width=20, text="Withdrawal", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=withdraw).place(x=20 ,y=320)
    Button(new, width=20, text="Transfer", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=transfer).place(x=20 ,y=410)
    Button(new, width=20, text="Change Pin", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=change).place(x=550 ,y=230)
    Button(new, width=20, text="Quickteller", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=quickteller).place(x=550 ,y=320)
    Button(new, width=20, text="Account Statement", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=statement).place(x=550 ,y=410)
    Button(new, width=20, text="Take Loan", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=loan).place(x=20 ,y=500)
    Button(new, width=20, text="Terminate Account", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=terminate).place(x=550 ,y=500)
    Button(new, width=20, text="Logout", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=logout).place(x=550 ,y=590)
    Button(new, width=20, text="Repay Debt", font=("arial", 25, "bold"),bg="#3697f5", bd=1, border=2, command=repay).place(x=20 ,y=590)


statements = {
    "atmospheric pressure":"1",
    "Bank Of Fuckery":"2",
    "This is a completely random message":"3",
    "Nameless Bank":"4",
    "I AM HIM":"5",
    "Your account is quite literally on RED":"6",
    "Account balance: $9999999999":"7",
    "How much money do you think you have\nYou are swimming in debt":"8"
}


choices = random.randint(1,8)
choice = str(choices)

for i in statements:
    if statements[i] == choice:
        dec=i
        break


root=Tk()
root.geometry("600x600")
root.title("ATM Machine")
root.resizable(False,False)
root.config(bg="black")

image_icon=PhotoImage(file="Py_Project Directory\Tesla.png")   # type: ignore
root.iconphoto(False, image_icon)

imag=ImageTk.PhotoImage(Image.open("Py_Project Directory\\Untitled59.jpg"))   # type: ignore
photo=Label(image=imag, bg="black")
photo.pack()

labels = Label(root, text=f"{dec}", width=50, font=("arial", 20, "bold"),bg="black", fg="red")     # type: ignore
labels.place(x=-120, y=10)

label = Label(root ,text="Enter Username", width=15, font=("arial", 23),  bg="black", fg="red")
label.place(x=10, y=350)  

name=StringVar()
Entry(width=12, font=("arial", 28), textvariable=name, bd=0, border=2).place(x=20, y=410)

label = Label(root ,text="Enter Passcode", width=15, font=("arial", 23), bg="black", fg="red")
label.place(x=310, y=350) 

code=StringVar()
Entry(width=12, font=("arial", 28), textvariable=code, bd=0,show="*", border=2).place(x=320, y=410)

def author():
    user = name.get() 
    users = str(user)
    codes = code.get()
    coder = int(codes)

    run = 1

    for i in range(2, ws_1.max_row+1):
        b = ws_1.cell(row=i, column=3)
        bs = b.value
        if users != str(bs):
            run+=1
            if run == ws_1.max_row:
                messagebox.showinfo("ERROR", "Username not Found")
        elif users == str(bs):  # type: ignore
            pw = ws_1.cell(row=i, column=2)
            pws = pw.value
            
    if str(coder) == str(pws):  # type: ignore
        logpage()
    else:
        messagebox.showinfo("INFO", "Wrong Passkey")


Button(root, width=10, text="Login", font=("arial", 20, "bold"),bg="#3697f5", bd=1, border=2, command=author).place(x=200 ,y=500)

root.mainloop()